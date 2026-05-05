# -*- coding: utf-8 -*-
"""
1) Učitava xlsx (kolone Ime, Sektor, Opis radnog mesta).
2) Učitava trenutna imena iz fajla scripts/_db_full_names.txt (ili drugi put kao argv[2]).
3) Generiše SQL: UPDATE preko full_name + INSERT za nove redove.
"""
import json
import re
import sys
import unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
ALIASES_PATH = ROOT / "scripts" / "payroll_name_aliases.json"
# Redovi u Excelu koji nisu zaposleni (naslovi, sekcije, greške)
SKIP_IME = frozenset(
    s.strip().lower()
    for s in [
        "HAP FLUID",
        "HAP Fluid",
        "Ime",
        "NEBOJŠA HPA",
        "Nebojša HPA",
        "Dragoslav",
        # agregat / praksa / placeholder (nisu zaposleni u employees)
        "OSTALI",
        "PRAKSA",
        "Nedjo",
        "Nedžo",
        "Vladimir",
        "Rajko",
        "STRANCI",
        "DUALNO",
    ]
)


def strip_d(s: str) -> str:
    if not s:
        return ""
    n = unicodedata.normalize("NFD", s)
    return "".join(c for c in n if unicodedata.category(c) != "Mn")


def norm_key(name: str) -> str:
    s = strip_d(name.lower())
    s = re.sub(r"[.,'\"]+", " ", s)
    parts = [p for p in s.split() if p]
    return " ".join(sorted(parts))


def to_db_surname_first(excel_ime: str) -> str:
    parts = excel_ime.strip().split()
    if len(parts) < 2:
        return excel_ime.strip()
    return f"{parts[-1]} {' '.join(parts[:-1])}"


def map_department(sektor: str, opis: str) -> str:
    s = strip_d((sektor or "").lower())
    o = strip_d((opis or "").lower())
    if "tehnolog" in o or o == "tehnogija":
        return "Tehnologija"
    if "zavar" in o:
        return "Bravarija"
    if "sečenj" in o or "secen" in o or "rezan" in o:
        return "Sečenje"
    if any(
        x in o
        for x in (
            "strug",
            "cnc",
            "borverk",
            "obaranje",
        )
    ) or (("operater" in o) and ("maš" in o or "mas" in o)):
        return "Mašinska obrada"
    if "operater" in o and "ma" in o:
        return "Mašinska obrada"
    if "farbar" in o or "priprem" in o or "peskare" in o:
        return "Farbara"
    if "planir" in o or o.startswith("planer"):
        return "Planiranje"
    if any(
        x in o
        for x in (
            "logist",
            "transport",
            "vozač",
            "vozac",
            "vozni",
            "viljusk",
        )
    ):
        return "Logistika"
    if "magacin" in o or "magacioner" in o:
        return "Magacin"
    if "nabavk" in o:
        return "Nabavka"
    if "portir" in o:
        return "Portirnica"
    if "bzr" in o or "bezbednost" in o:
        return "BZR"
    if "kooperac" in o:
        return "Kooperacija"
    if "monta" in o and "ma" in o:
        return "Mašinska montaža"
    if "održavan" in o or "odrzavan" in o or "servis" in o or "tehničar elektro" in o:
        return "Održavanje"
    if "elektro" in o and "mont" not in o and "odr" not in o:
        return "Elektro projektovanje"
    if "elektro" in o:
        return "Elektro montaža"
    if "kontrol" in o and "kvalit" in o:
        return "Kontrola kvaliteta"
    if s == "proizvodnja":
        return "Proizvodnja"
    if s == "logistika":
        return "Logistika"
    if s:
        return (sektor or "").strip()
    return "Servoteh"


def normalize_department(dep: str) -> str:
    d = (dep or "").strip()
    m = {
        "Kontrola": "Kontrola kvaliteta",
        "Menadzment": "Menadžment",
        "Montaza": "Mašinska montaža",
        "montaza": "Mašinska montaža",
        # isključivo "Elektro" (ne projektovanje) → montaža, kao u map_department
        "Elektro": "Elektro montaža",
        "Inzenjer prodaje": "Inženjer prodaje",
    }
    return m.get(d, d)


def esc_sql(s: str) -> str:
    return "'" + str(s).replace("'", "''") + "'"


def main() -> None:
    xlsx_path = (
        sys.argv[1]
        if len(sys.argv) > 1
        else r"Z:\Obracun\3. Isplata sati za MART 2026.xlsx"
    )
    names_path = (
        Path(sys.argv[2])
        if len(sys.argv) > 2
        else ROOT / "scripts" / "_db_full_names.txt"
    )
    aliases = {}
    if ALIASES_PATH.is_file():
        aliases = json.loads(ALIASES_PATH.read_text(encoding="utf-8"))
    with open(names_path, "r", encoding="utf-8") as f:
        db_names = [ln.strip() for ln in f if ln.strip()]

    by_norm: dict = {}
    for n in db_names:
        by_norm[norm_key(n)] = n

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = "JUN 2025" if "JUN 2025" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]
    row1 = [c.value for c in ws[1]]

    def cidx(hname: str) -> int:
        return row1.index(hname)

    i_ime = cidx("Ime")
    try:
        i_sek = cidx("Sektor")
    except ValueError:
        i_sek = None
    try:
        i_opis = cidx("Opis radnog mesta")
    except ValueError:
        i_opis = None

    excel_rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        raw = row[i_ime]
        if not raw or not str(raw).strip():
            continue
        excel_ime = str(raw).strip()
        if not excel_ime or excel_ime.strip().lower() in SKIP_IME:
            continue
        sektor_s = str(row[i_sek]).strip() if i_sek is not None and row[i_sek] else ""
        opis_s = str(row[i_opis]).strip() if i_opis is not None and row[i_opis] else ""
        dep = (
            normalize_department(map_department(sektor_s, opis_s))
            if (sektor_s or opis_s)
            else "Servoteh"
        )
        excel_rows.append(
            {
                "excel_ime": excel_ime,
                "db_name": to_db_surname_first(excel_ime),
                "norm": norm_key(excel_ime),
                "sektor": sektor_s,
                "opis": opis_s,
                "department": dep,
            }
        )

    def resolve_alias(excel_ime: str) -> str | None:
        for k, v in aliases.items():
            if k.strip().lower() == excel_ime.strip().lower():
                return v
        return None

    updates = []
    inserts = []
    for er in excel_rows:
        alias_fn = resolve_alias(er["excel_ime"])
        if alias_fn:
            hit = by_norm.get(norm_key(alias_fn))
        else:
            hit = by_norm.get(er["norm"]) or by_norm.get(norm_key(er["db_name"]))
        if hit:
            new_pos = (er["opis"] or "")[:500]
            updates.append(
                {
                    "full_name": hit,
                    "new_dep": er["department"],
                    "new_pos": new_pos,
                }
            )
        else:
            inserts.append(er)

    out_sql: list = []
    out_sql.append(
        f"-- employees sync from payroll: sheet {sheet} ({len(excel_rows)} redova u Excelu)\n"
    )
    out_sql.append("BEGIN;\n\n")
    for u in updates:
        out_sql.append(
            f"UPDATE employees SET department = {esc_sql(u['new_dep'])}, position = {esc_sql(u['new_pos'])} "
            f"WHERE full_name = {esc_sql(u['full_name'])};\n"
        )
    for ins in inserts:
        out_sql.append(
            f"INSERT INTO employees (full_name, department, position, is_active) VALUES ("
            f"{esc_sql(ins['db_name'])}, {esc_sql(ins['department'])}, {esc_sql((ins['opis'] or '')[:500])}, true); "
            f"-- excel: {ins['excel_ime']}\n"
        )
    out_sql.append("\nCOMMIT;\n")

    sql_path = ROOT / "sql" / "migrations" / "sync_employees_from_payroll_mart_2026.sql"
    sql_path.write_text("".join(out_sql), encoding="utf-8")

    rep = {
        "sheet": sheet,
        "excel_employee_rows": len(excel_rows),
        "db_name_lines": len(db_names),
        "update_statements": len(updates),
        "insert_statements": len(inserts),
    }
    (ROOT / "scripts" / "_payroll_sync_report.json").write_text(
        json.dumps(
            {**rep, "insert_preview": [x["excel_ime"] for x in inserts[:50]]},
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(json.dumps(rep, ensure_ascii=False, indent=2))
    print("Wrote", sql_path, file=sys.stderr)


if __name__ == "__main__":
    main()
