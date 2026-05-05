"""
Export Kadrovska tabele public.employees u CSV, HTML i XLSX.

Izvor: MCP execute_sql JSON fajl, ili --from-csv na postojeci employees_full_export.csv.

XLSX: pip install -r scripts/requirements-kadrovska-export.txt
"""
from __future__ import annotations

import argparse
import csv
import html
import json
import sys
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "docs" / "reports"
SCHEMA_CSV = OUT_DIR / "employees_table_schema.csv"
DATA_CSV = OUT_DIR / "employees_full_export.csv"
SUMMARY_CSV = OUT_DIR / "employees_null_summary.csv"
REPORT_HTML = OUT_DIR / "employees_kadrovska_pregled.html"
DATA_XLSX = OUT_DIR / "employees_kadrovska_export.xlsx"
SCHEMA_DETAIL = OUT_DIR / "employees_table_schema_detail.csv"


def extract_json_rows(raw: str) -> list[dict]:
    """
    Parse MCP `execute_sql` text dump:
    {"result": "...<untrusted-data-...>\\n[{...}] or wrapped [{\"rows\": [...]}] ..." }
    """
    # 1) Outer JSON wrapper
    try:
        outer = json.loads(raw)
    except json.JSONDecodeError as e:
        raise ValueError(f"Not valid outer JSON: {e}") from e
    if not isinstance(outer, dict) or "result" not in outer:
        raise ValueError("Expected top-level {result: ...}")
    result_str = outer["result"]
    if not isinstance(result_str, str):
        raise ValueError("result is not a string")

    # 2) After untrusted block: JSON from first '[' that starts rows or a plain array
    if '"rows"' in result_str:
        i = result_str.find('"rows"')
        i = result_str.find("[", i)
    else:
        i = result_str.find("[{")
    if i == -1:
        raise ValueError("No JSON array in result string")

    decoder = json.JSONDecoder()
    data, _end = decoder.raw_decode(result_str, i)

    # 3) Unwrap common shapes
    if isinstance(data, list) and data and isinstance(data[0], dict) and "rows" in data[0]:
        inner = data[0]["rows"]
        if isinstance(inner, list):
            return inner
    if (
        isinstance(data, list)
        and len(data) == 1
        and isinstance(data[0], dict)
        and "data" in data[0]
    ):
        inner = data[0]["data"]
        if isinstance(inner, list):
            return inner
    if isinstance(data, list) and data and "full_name" in data[0]:
        return data
    if isinstance(data, dict) and "rows" in data and isinstance(data["rows"], list):
        return data["rows"]
    if isinstance(data, dict) and "data" in data and isinstance(data["data"], list):
        return data["data"]

    raise ValueError("Could not unwrap employee rows from parsed JSON")


def load_rows_from_csv(path: Path) -> list[dict]:
    rows: list[dict] = []
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if "is_active" in row and row["is_active"] is not None:
                t = (row["is_active"] or "").strip().lower()
                if t in ("true", "1", "yes", "da"):
                    row["is_active"] = True
                elif t in ("false", "0", "no", "ne"):
                    row["is_active"] = False
            rows.append(row)
    return rows


def fieldnames_for(rows: list[dict]) -> list[str]:
    all_keys: set[str] = set()
    for r in rows:
        all_keys.update(r.keys())
    preferred = [
        "id",
        "full_name",
        "first_name",
        "last_name",
        "department",
        "position",
        "phone",
        "email",
        "hire_date",
        "is_active",
        "note",
        "personal_id",
        "birth_date",
        "gender",
        "address",
        "city",
        "postal_code",
        "bank_name",
        "bank_account",
        "phone_private",
        "emergency_contact_name",
        "emergency_contact_phone",
        "slava",
        "slava_day",
        "education_level",
        "education_title",
        "medical_exam_date",
        "medical_exam_expires",
        "team",
        "created_at",
        "updated_at",
    ]
    return [k for k in preferred if k in all_keys] + sorted(
        k for k in all_keys if k not in preferred
    )


def is_value_empty_for_summary(col: str, v) -> bool:
    if col == "is_active":
        return v is None
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def build_summary(
    rows: list[dict], fieldnames: list[str]
) -> list[dict[str, str | int | float]]:
    n = len(rows)
    out: list[dict[str, str | int | float]] = []
    for col in fieldnames:
        empty = sum(1 for r in rows if is_value_empty_for_summary(col, r.get(col)))
        out.append(
            {
                "column": col,
                "empty_or_null": empty,
                "total_rows": n,
                "fill_pct": round(100.0 * (n - empty) / n, 1) if n else 0.0,
            }
        )
    return out


def write_xlsx(
    path: Path,
    rows: list[dict],
    fieldnames: list[str],
    summary: list[dict],
    by_dept: list[tuple[str, int]],
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise SystemExit(
            "openpyxl nije instaliran. Pokreni:\n"
            "  pip install -r scripts/requirements-kadrovska-export.txt"
        ) from e

    wb = Workbook()
    # Radnici
    ws1 = wb.active
    ws1.title = "Radnici"
    sorted_rows = sorted(
        rows,
        key=lambda x: (x.get("department") or "", x.get("full_name") or ""),
    )
    ws1.append(fieldnames)
    for r in sorted_rows:
        ws1.append([r.get(c) for c in fieldnames])
    ws1.freeze_panes = "A2"
    for i, c in enumerate(fieldnames, 1):
        if c in ("note", "address"):
            ws1.column_dimensions[get_column_letter(i)].width = 36
        else:
            ws1.column_dimensions[get_column_letter(i)].width = 16

    # Odeljenja
    ws2 = wb.create_sheet("Odeljenja")
    ws2.append(["Odeljenje", "Broj"])
    for dept, cnt in by_dept:
        ws2.append([dept, cnt])

    # Praznine
    ws3 = wb.create_sheet("Praznine")
    ws3.append(["Kolona", "Prazno", "Ukupno", "Popunjenost %"])
    for s in summary:
        ws3.append(
            [s["column"], s["empty_or_null"], s["total_rows"], s["fill_pct"]]
        )

    # Sema
    ws4 = wb.create_sheet("Sema")
    if SCHEMA_DETAIL.is_file():
        with SCHEMA_DETAIL.open(newline="", encoding="utf-8-sig") as f:
            rdr = csv.DictReader(f)
            h = rdr.fieldnames or []
            if h:
                ws4.append(list(h))
            for r in rdr:
                ws4.append([r.get(c, "") for c in h])
    else:
        ws4.append(["column_name"])
        for c in fieldnames:
            ws4.append([c])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def run_export(rows: list[dict], write_xlsx_file: bool) -> int:
    for r in rows:
        r.pop("sort_dept", None)
        r.pop("sort_name", None)

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    fieldnames = fieldnames_for(rows)
    n = len(rows)
    summary_rows = build_summary(rows, fieldnames)

    with DATA_CSV.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        for r in sorted(
            rows,
            key=lambda x: (x.get("department") or "", x.get("full_name") or ""),
        ):
            w.writerow(r)

    with SUMMARY_CSV.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(
            f,
            fieldnames=["column", "empty_or_null", "total_rows", "fill_pct"],
        )
        w.writeheader()
        for s in summary_rows:
            w.writerow(s)

    # Department counts
    by_dept = Counter()
    for r in rows:
        by_dept[(r.get("department") or "").strip() or "(prazno)"] += 1
    by_dept_list = sorted(by_dept.items(), key=lambda x: (-x[1], x[0]))
    if write_xlsx_file:
        write_xlsx(
            DATA_XLSX, rows, fieldnames, summary_rows, by_dept_list
        )

    # HTML
    def esc(s: str) -> str:
        return html.escape(s or "")

    dept_rows = "".join(
        f"<tr><td>{esc(d)}</td><td class=\"n\">{c}</td></tr>"
        for d, c in sorted(by_dept.items(), key=lambda x: (-x[1], x[0]))
    )
    # Table: show key columns in main, full data in scroll
    key_cols = [
        "full_name",
        "department",
        "position",
        "phone",
        "email",
        "hire_date",
        "is_active",
        "personal_id",
        "city",
    ]
    key_cols = [c for c in key_cols if c in fieldnames]

    thead = "<tr>" + "".join(f"<th>{esc(c)}</th>" for c in key_cols) + "</tr>"
    body_parts = []
    for r in sorted(
        rows,
        key=lambda x: (x.get("department") or "", x.get("full_name") or ""),
    ):
        tds = []
        for c in key_cols:
            v = r.get(c)
            if v is None:
                tds.append('<td class="empty">—</td>')
            elif isinstance(v, bool):
                tds.append(f"<td>{'da' if v else 'ne'}</td>")
            else:
                tds.append(f"<td>{esc(str(v))}</td>")
        body_parts.append("<tr>" + "".join(tds) + "</tr>")
    table_main = "\n".join(body_parts)

    # Full table all columns
    thead_full = "<tr>" + "".join(f"<th>{esc(c)}</th>" for c in fieldnames) + "</tr>"
    full_rows = []
    for r in sorted(
        rows,
        key=lambda x: (x.get("department") or "", x.get("full_name") or ""),
    ):
        tds = []
        for c in fieldnames:
            v = r.get(c)
            if v is None or v == "":
                tds.append('<td class="empty">—</td>')
            else:
                tds.append(f"<td>{esc(str(v))}</td>")
        full_rows.append("<tr>" + "".join(tds) + "</tr>")
    table_full = "\n".join(full_rows)

    html_out = f"""<!DOCTYPE html>
<html lang="sr">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>Kadrovska — employees pregled</title>
<style>
  body {{ font-family: Segoe UI, system-ui, sans-serif; margin: 24px; color: #1a1a1a; line-height: 1.4; }}
  h1 {{ font-size: 1.35rem; font-weight: 600; }}
  h2 {{ font-size: 1.1rem; margin-top: 2rem; font-weight: 600; }}
  p.meta {{ color: #555; font-size: 0.9rem; }}
  table {{ border-collapse: collapse; width: 100%; font-size: 0.85rem; }}
  th, td {{ border: 1px solid #ccc; padding: 6px 8px; text-align: left; vertical-align: top; }}
  th {{ background: #f0f0f0; position: sticky; top: 0; z-index: 1; }}
  .n {{ text-align: right; }}
  .empty {{ color: #999; }}
  .wrap {{ overflow-x: auto; max-width: 100%; border: 1px solid #ddd; }}
  .note {{ background: #f8f8f8; padding: 12px; border-left: 3px solid #333; margin: 1rem 0; }}
</style>
</head>
<body>
<h1>Kadrovska — tabela <code>public.employees</code></h1>
<p class="meta">Izvoz: {n} zaposlenih. XLSX: <code>docs/reports/employees_kadrovska_export.xlsx</code> (listovi: Radnici, Odeljenja, Praznine, Sema). CSV: <code>docs/reports/employees_full_export.csv</code> (UTF-8 sa BOM). Šema kolona: <code>docs/reports/employees_table_schema.csv</code>.</p>
<div class="note">
  <strong>Šta gledati:</strong> kolone sa niskim <em>fill %</em> u pregledu praznina (CSV <code>employees_null_summary.csv</code>) — tu najčešće fale podaci ili nisu uneti u UI.
</div>

<h2>Broj zaposlenih po odeljenju</h2>
<div class="wrap">
<table>
<thead><tr><th>Odeljenje</th><th>Broj</th></tr></thead>
<tbody>
{dept_rows}
</tbody>
</table>
</div>

<h2>Skraćeni pregled (ključne kolone)</h2>
<div class="wrap">
<table>
<thead>{thead}</thead>
<tbody>
{table_main}
</tbody>
</table>
</div>

<h2>Sve kolone (puni prikaz)</h2>
<div class="wrap">
<table>
<thead>{thead_full}</thead>
<tbody>
{table_full}
</tbody>
</table>
</div>
</body>
</html>
"""
    REPORT_HTML.write_text(html_out, encoding="utf-8")

    # schema stub from static list (31 cols from live DB) — user can diff
    SCHEMA_CSV.write_text(
        "column_name\n"
        + "\n".join(fieldnames),
        encoding="utf-8-sig",
    )

    print(f"Wrote {DATA_CSV}")
    print(f"Wrote {SUMMARY_CSV}")
    print(f"Wrote {REPORT_HTML}")
    print(f"Wrote {SCHEMA_CSV}")
    if write_xlsx_file:
        print(f"Wrote {DATA_XLSX}")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Izvoz public.employees u CSV, HTML, XLSX (openpyxl)."
    )
    ap.add_argument(
        "mcp_file",
        nargs="?",
        help="Fajl sa MCP execute_sql JSON odgovorom (wrapper sa kljucem 'result')",
    )
    ap.add_argument(
        "--from-csv",
        type=Path,
        metavar="PATH",
        help="Ucitaj redove iz postojeceg employees_full_export.csv (bez MCP).",
    )
    ap.add_argument(
        "--no-xlsx",
        action="store_true",
        help="Ne generisi employees_kadrovska_export.xlsx",
    )
    args = ap.parse_args()
    if args.mcp_file and args.from_csv:
        print("Koristi ili mcp_fajl ILI --from-csv, ne oba.", file=sys.stderr)
        return 1
    if not args.mcp_file and not args.from_csv:
        ap.print_help()
        return 1

    if args.from_csv:
        rows = load_rows_from_csv(args.from_csv)
    else:
        raw = Path(args.mcp_file).read_text(encoding="utf-8", errors="replace")
        rows = extract_json_rows(raw)
    if not rows:
        print("Nema redova.", file=sys.stderr)
        return 1

    return run_export(rows, write_xlsx_file=not args.no_xlsx)


if __name__ == "__main__":
    raise SystemExit(main())
