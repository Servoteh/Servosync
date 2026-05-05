"""
Čita docs/reports/employees_kadrovska_export.xlsx (list Radnici) i ispisuje SQL
UPDATE naredbe za public.employees (id iz kolone; created_at se ne menja;
updated_at = now()).
"""
from __future__ import annotations

import re
import sys
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any
from uuid import UUID

import openpyxl
from openpyxl.cell.cell import Cell

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "docs" / "reports" / "employees_kadrovska_export.xlsx"

# id: samo WHERE; created_at: ne diramo; updated_at: uvek now() u bazi, ne iz tabele
# full_name, first_name, last_name: ne iz ćelija (Excel puca na š/ć) — vidi split_prezime_ime
SKIP_COLS = {"id", "created_at", "updated_at", "full_name", "first_name", "last_name"}


def split_prezime_ime(fn: str | None) -> tuple[str | None, str | None, str | None]:
    """
    Isto pravilo kao u bazi: prva reč = prezime, ostatak = ime.
    Prazan string -> tri NULL.
    Jedna reč -> samo last_name, first NULL.
    """
    if fn is None or not str(fn).strip():
        return None, None, None
    canon = re.sub(r"\s+", " ", str(fn).strip())
    if not canon:
        return None, None, None
    sp = canon.find(" ")
    if sp == -1:
        return canon, None, canon
    last = canon[:sp]
    first = (canon[sp + 1 :].strip() or None) if sp >= 0 else None
    return canon, first, last


def sql_literal(v: Any) -> str:
    if v is None or v == "":
        return "NULL"
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float, Decimal)) and not isinstance(v, bool):
        return str(v)
    if isinstance(v, datetime):
        return f"'{v.isoformat()}'::timestamptz"
    if isinstance(v, date) and not isinstance(v, datetime):
        return f"'{v.isoformat()}'::date"
    if isinstance(v, time):
        return f"'{v.isoformat()}'::time"
    if isinstance(v, UUID):
        return f"'{v}'::uuid"
    s = str(v).strip()
    if s == "":
        return "NULL"
    return "'" + s.replace("'", "''") + "'"


def cell_value(c: Any) -> Any:
    if c is None:
        return None
    if isinstance(c, Cell) and c.value is None and c.data_type == "e":
        return None
    v = c.value if isinstance(c, Cell) else c
    if isinstance(v, str):
        t = v.strip()
        if t in ("#N/A", "#REF!", "#VALUE!"):
            return None
        return t if t else None
    return v


def main() -> int:
    out_path: Path | None = None
    args = sys.argv[1:]
    if "--out" in args:
        i = args.index("--out")
        out_path = Path(args[i + 1])
        args = args[:i] + args[i + 2 :]
    xlsx = Path(args[0]) if args else DEFAULT_XLSX
    if not xlsx.is_file():
        print(f"Nema fajla: {xlsx}", file=sys.stderr)
        return 1
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if "Radnici" not in wb.sheetnames:
        print(f"Nema lista Radnici: {wb.sheetnames}", file=sys.stderr)
        return 1
    ws = wb["Radnici"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("Prazan sheet", file=sys.stderr)
        return 1
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx = {name: i for i, name in enumerate(header) if name}
    if "id" not in idx:
        print("Nema kolone id", file=sys.stderr)
        return 1
    out_lines: list[str] = [
        "BEGIN;",
        "SET statement_timeout = '120s';",
    ]
    for r in rows[1:]:
        if r is None or all(x is None or str(x).strip() == "" for x in r):
            continue
        row_id = cell_value(r[idx["id"]])
        if not row_id:
            continue
        fn_raw = cell_value(r[idx["full_name"]]) if "full_name" in idx else None
        canon, first_n, last_n = split_prezime_ime(
            str(fn_raw) if fn_raw is not None else None
        )
        if canon is None:
            print(f"Preskačem red bez full_name: id={row_id}", file=sys.stderr)
            continue
        sets: list[str] = [
            f"  full_name = {sql_literal(canon)}",
            f"  first_name = {sql_literal(first_n)}",
            f"  last_name = {sql_literal(last_n)}",
        ]
        for col in header:
            if not col or col in SKIP_COLS or col not in idx:
                continue
            v = cell_value(r[idx[col]])
            if col == "is_active" and v is not None and isinstance(v, str):
                t = v.strip().lower()
                if t in ("true", "1", "da", "yes", "d"):
                    v = True
                elif t in ("false", "0", "ne", "no", "n"):
                    v = False
            if col in (
                "hire_date",
                "birth_date",
                "medical_exam_date",
                "medical_exam_expires",
            ) and isinstance(v, datetime):
                v = v.date()
            sets.append(f'  {col} = {sql_literal(v)}')
        sets.append("  updated_at = now()")
        out_lines.append(
            f"UPDATE public.employees\nSET\n{',\n'.join(sets)}\nWHERE id = {sql_literal(str(row_id))}::uuid;\n"
        )
    out_lines.append("COMMIT;")
    text = "\n".join(out_lines) + "\n"
    if out_path is not None:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(text, encoding="utf-8")
    else:
        sys.stdout.buffer.write(text.encode("utf-8"))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
