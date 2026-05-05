"""Ispravka poznatih oštećenja u employees_kadrovska_export.xlsx (š u imenu, slučajni razmaci)."""
from __future__ import annotations

import openpyxl
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "docs" / "reports" / "employees_kadrovska_export.xlsx"

FIXES_BY_ID: dict[str, dict[str, str | None]] = {
    "00b5d49e-24ea-4da4-8a34-386c0ee4d5b0": {
        "first_name": "Aleksandar",
    },
    "0764e73d-fa40-4239-901e-a9a4c6586790": {
        "first_name": "Branislava",
    },
    "698417f5-cf88-4ae6-97c6-0ab862ceadfa": {
        "first_name": "Aleksandar",
    },
    "982d732e-9d3b-4b75-b845-efbcad0c3d3d": {
        "full_name": "Kaštratović Dijana",
        "first_name": "Dijana",
        "last_name": "Kaštratović",
    },
}


def main() -> int:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Radnici"]
    header = [c.value for c in ws[1]]
    col = {n: header.index(n) for n in header if n}
    for r in range(2, ws.max_row + 1):
        eid = ws.cell(r, col["id"] + 1).value
        if not eid:
            continue
        sid = str(eid)
        if sid not in FIXES_BY_ID:
            continue
        for k, v in FIXES_BY_ID[sid].items():
            if k in col:
                ws.cell(r, col[k] + 1, value=v)
    wb.save(XLSX)
    print(f"Saved {XLSX}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
